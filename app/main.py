from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from typing import Dict, List, Any
from pathlib import Path
import re, uuid
from openpyxl import load_workbook
from pydantic import BaseModel
from pydantic import BaseModel, Field

# =============================================================================
# App setup
# =============================================================================
app = FastAPI(title="AI Maths Prep API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

APP_DIR = Path(__file__).parent
TUTORIALS_DIR = APP_DIR / "data" / "tutorials"
QUIZZES_DIR = APP_DIR / "data" / "quizzes"

# serve static tutorials
app.mount("/static", StaticFiles(directory="app/data"), name="static")

# =============================================================================
# Health check
# =============================================================================
@app.get("/healthz")
def healthz():
    return {"status": "ok"}

# =============================================================================
# Tutorials
# =============================================================================
@app.get("/tutorials_list")
def tutorials_list(subject: str):
    subject_dir = TUTORIALS_DIR / subject
    if not subject_dir.exists():
        raise HTTPException(status_code=404, detail=f"Subject not found: {subject}")
    files = sorted([p.name for p in subject_dir.glob("*.pdf")])
    return [{"filename": f, "title": f[:-4].replace("_"," ").title(), "subject": subject} for f in files]

@app.get("/tutorials_file")
def tutorials_file(subject: str, filename: str):
    file_path = TUTORIALS_DIR / subject / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Tutorial file not found")
    return {"url": f"/data/tutorials/{subject}/{filename}"}

# =============================================================================
# Quiz Engine
# =============================================================================
class QuizStart(BaseModel):
    subject: str = Field(..., example="LA")
    topic: str = Field(..., example="1.1")


class QuizAnswer(BaseModel):
    session_id: str = Field(..., example="123e4567-e89b-12d3-a456-426614174000")
    answer: int = Field(..., example=2)

    class Config:
        schema_extra = {
            "example": {
                "session_id": "123e4567-e89b-12d3-a456-426614174000",
                "answer": 2
            }
        }


_SESSIONS: Dict[str, Dict[str, Any]] = {}

def _sheet_key(topic: str) -> str:
    m = re.search(r"(\d+)\.(\d+)", topic)
    return f"{m.group(1)}.{m.group(2)}" if m else topic

def _load_questions(subject: str, sheet: str) -> List[Dict[str, Any]]:
    xlsx = QUIZZES_DIR / f"{subject}.xlsx"
    if not xlsx.exists():
        raise HTTPException(status_code=400, detail=f"Quiz workbook missing: {xlsx}")
    wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise HTTPException(status_code=400, detail=f"Sheet {sheet} not in workbook")
    ws = wb[sheet]
    hdr = [str(c).strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {c: i for i, c in enumerate(hdr)}

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["question text"]]: continue
        opts = [row[idx["option a"]], row[idx["option b"]], row[idx["option c"]], row[idx["option d"]]]
        opts = ["" if o is None else str(o) for o in opts]
        corr = str(row[idx["correct answer"]]).strip().upper()
        corr_idx = {"A":0,"B":1,"C":2,"D":3}.get(corr, 0)
        diff = str(row[idx["difficulty"]] or "").lower()
        if diff.startswith("e"): diff="easy"
        elif diff.startswith("m"): diff="medium"
        else: diff="hard"
        out.append({"text": row[idx["question text"]], "options": opts, "correct": corr_idx, "difficulty": diff})
    wb.close()
    return out

def _classification(score: int) -> str:
    if score <= 3: return "Beginner"
    if score <= 6: return "Confident"
    if score <= 9: return "Master"
    return "Champion"

def _shift(diff: str, correct: bool) -> str:
    levels = ["easy","medium","hard"]
    i = levels.index(diff)
    if correct and i < 2: i += 1
    if not correct and i > 0: i -= 1
    return levels[i]

@app.post("/quiz/start")
def quiz_start(req: QuizStart):
    sheet = _sheet_key(req.topic)
    all_qs = _load_questions(req.subject, sheet)

    pools = {"easy":[],"medium":[],"hard":[]}
    for q in all_qs: pools[q["difficulty"]].append(q)

    sid = str(uuid.uuid4())
    _SESSIONS[sid] = {
        "subject": req.subject, "sheet": sheet,
        "pools": pools,
        "step": 1, "score": 0, "p1_score": 0,
        "consec_correct": 0, "current": None
    }
    # Q1 = Easy if available
    q = pools["easy"].pop(0) if pools["easy"] else pools["medium"].pop(0)
    _SESSIONS[sid]["current"] = q
    return {"session_id": sid, "question": {"text": q["text"], "options": q["options"]}, "done": False}

@app.post("/quiz/answer")
def quiz_answer(req: QuizAnswer):
    s = _SESSIONS.get(req.session_id)
    if not s: raise HTTPException(status_code=400, detail="Invalid session_id")
    q = s["current"]; 
    if not q: raise HTTPException(status_code=400, detail="No current question")

    correct = (req.answer == q["correct"])
    if correct:
        s["score"] += 1; s["consec_correct"] += 1
    else:
        s["consec_correct"] = 0

    # ---------------- Phase 1 (Q1-5) ----------------
    if 1 <= s["step"] < 5:
        if correct: s["p1_score"] += 1
        s["step"] += 1
        next_diff = "easy" if s["step"] <= 2 else "medium"
        nq = s["pools"][next_diff].pop(0)
        s["current"] = nq
        return {"correct": correct, "done": False, "question": {"text": nq["text"], "options": nq["options"]}}

    if s["step"] == 5:
        if correct: s["p1_score"] += 1
        if s["p1_score"] <= 2:
            return {"correct": correct, "done": True, "total": 5, "score": s["score"],
                    "unlock_next": False,
                    "status": "Poor understanding. Review again.",
                    "classification": _classification(s["score"])}
        if s["p1_score"] in [3,4]:
            return {"correct": correct, "done": True, "total": 5, "score": s["score"],
                    "unlock_next": True,
                    "status": "You're ready for next tutorial",
                    "classification": _classification(s["score"])}
        # Perfect → continue Phase 2
        s["step"] = 6
        nq = s["pools"]["hard"].pop(0)
        s["current"] = nq
        return {"correct": correct, "done": False, "question": {"text": nq["text"], "options": nq["options"]}}

    # ---------------- Phase 2 (Q6-Q12) ----------------
    if 6 <= s["step"] <= 12:
        if s["consec_correct"] >= 3 or s["step"] == 12:
            return {"correct": correct, "done": True, "total": s["step"], "score": s["score"],
                    "unlock_next": True, "status": "Quiz finished",
                    "classification": _classification(s["score"])}
        s["step"] += 1
        next_diff = _shift(q["difficulty"], correct)
        nq = s["pools"][next_diff].pop(0)
        s["current"] = nq
        return {"correct": correct, "done": False, "question": {"text": nq["text"], "options": nq["options"]}}
